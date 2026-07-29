"""
Layer 1: Excel Sheet 自动检测与解析器

功能:
- 自动识别实体 Sheet 和关系 Sheet（不依赖硬编码名称）
- 支持中英文多种变体表头
- 返回标准化的 ParsedSheet 结构
"""

import io
from dataclasses import dataclass, field
from typing import Optional

from openpyxl import load_workbook


# ---------------------------------------------------------------------------
# 关键词定义（支持中英文多种变体）
# ---------------------------------------------------------------------------

ENTITY_LABEL_KEYWORDS = [
    "标签", "label", "type", "类型", "entity_type", "类别", "category",
    "节点标签", "node_label", "节点类型",
]
ENTITY_NAME_KEYWORDS = [
    "名称", "name", "实体名称", "node_name", "实体", "entity",
    "节点名称", "entity_name", "节点名",
]

REL_SOURCE_KEYWORDS = [
    "源实体", "source", "src", "from", "start", "源", "source_name",
    "源名称", "起始实体", "起始节点", "source_entity",
]
REL_TYPE_KEYWORDS = [
    "关系类型", "type", "relation", "rel_type", "关系", "关系名",
    "relationship", "edge_type", "边类型", "关系名称",
]
REL_TARGET_KEYWORDS = [
    "目标实体", "target", "to", "end", "dst", "目标", "target_name",
    "目标名称", "终止实体", "终止节点", "target_entity",
]

# 需要跳过的引导 Sheet 名称关键词
SKIP_SHEET_KEYWORDS = [
    "使用说明", "说明", "readme", "帮助", "help", "guide",
    "模板说明", "注意事项", "instructions", "introduction", "read",
]


# ---------------------------------------------------------------------------
# 数据结构
# ---------------------------------------------------------------------------

@dataclass
class ParsedSheet:
    """解析后的 Excel Sheet 数据"""
    sheet_name: str
    sheet_type: str                 # "entity" | "relationship" | "unknown"
    headers: list[str] = field(default_factory=list)  # 原始表头
    column_map: dict[str, int] = field(default_factory=dict)  # 规范化列名 → 列索引
    rows: list[dict] = field(default_factory=list)  # 每行数据 {规范化列名: 值}
    row_count: int = 0


@dataclass
class ExcelParseResult:
    """Excel 解析结果"""
    entity_sheet: Optional[ParsedSheet] = None
    relationship_sheet: Optional[ParsedSheet] = None
    unmatched_sheets: list[str] = field(default_factory=list)
    parse_errors: list[str] = field(default_factory=list)
    all_sheets: list[dict] = field(default_factory=list)  # [{name, type, headers, row_count}]


class ExcelSheetDetector:
    """Excel Sheet 自动检测与解析器"""

    # ------------------------------------------------------------------
    # 公开 API
    # ------------------------------------------------------------------

    def detect_sheets(self, file_bytes: bytes) -> ExcelParseResult:
        """检测并解析 Excel 中所有 Sheet。"""
        result = ExcelParseResult()

        try:
            wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        except Exception as e:
            result.parse_errors.append(f"无法打开 Excel 文件: {e}")
            return result

        sheet_scores: list[tuple[str, str, float]] = []  # (sheet_name, type, score)

        for sheet_name in wb.sheetnames:
            if self._should_skip_sheet(sheet_name):
                continue

            ws = wb[sheet_name]
            if ws.max_row < 2 or ws.max_column < 1:
                continue

            headers = self._read_headers(ws)
            if not headers:
                continue

            # 对每个 Sheet 计算实体分和关系分
            e_score = self._match_entity_score(headers)
            r_score = self._match_relationship_score(headers)

            cell_count = min(ws.max_row - 1, 100)  # 取样预览
            sheet_info = {
                "name": sheet_name,
                "type": "unknown",
                "headers": headers,
                "row_count": cell_count,
            }

            if e_score >= 0.6 and e_score > r_score:
                sheet_info["type"] = "entity"
                result.all_sheets.append(sheet_info)
                sheet_scores.append((sheet_name, "entity", e_score))
            elif r_score >= 0.6 and r_score >= e_score:
                sheet_info["type"] = "relationship"
                result.all_sheets.append(sheet_info)
                sheet_scores.append((sheet_name, "relationship", r_score))
            else:
                sheet_info["type"] = "unknown"
                result.all_sheets.append(sheet_info)

        # 取每类中得分最高的 Sheet
        entity_candidates = sorted(
            [(n, s) for n, t, s in sheet_scores if t == "entity"],
            key=lambda x: -x[1],
        )
        rel_candidates = sorted(
            [(n, s) for n, t, s in sheet_scores if t == "relationship"],
            key=lambda x: -x[1],
        )

        if entity_candidates:
            sheet_name = entity_candidates[0][0]
            result.entity_sheet = self._parse_sheet(wb[sheet_name], "entity")

        if rel_candidates:
            sheet_name = rel_candidates[0][0]
            result.relationship_sheet = self._parse_sheet(wb[sheet_name], "relationship")

        result.unmatched_sheets = [
            s["name"] for s in result.all_sheets if s["type"] == "unknown"
        ]

        wb.close()
        return result

    def preview_entity_sheet(self, file_bytes: bytes, sheet_name: str) -> Optional[ParsedSheet]:
        """手动指定 Sheet 名称预览实体数据。"""
        return self._parse_sheet_by_name(file_bytes, sheet_name, "entity")

    def preview_relationship_sheet(self, file_bytes: bytes, sheet_name: str) -> Optional[ParsedSheet]:
        """手动指定 Sheet 名称预览关系数据。"""
        return self._parse_sheet_by_name(file_bytes, sheet_name, "relationship")

    # ------------------------------------------------------------------
    # 内部方法
    # ------------------------------------------------------------------

    def _should_skip_sheet(self, name: str) -> bool:
        normalized = name.strip().lower()
        return any(kw in normalized for kw in SKIP_SHEET_KEYWORDS)

    def _read_headers(self, ws) -> list[str]:
        """读取第一行作为表头，惰性避免空元组。"""
        row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not row:
            return []
        return [self._normalize_header(str(v)) if v is not None else "" for v in row]

    @staticmethod
    def _normalize_header(raw: str) -> str:
        """对表头做归一化处理：去空格、去下划线、lower。用于匹配，但不改变显示值。"""
        return raw.strip().replace("_", "").lower()

    def _match_entity_score(self, headers: list[str]) -> float:
        labels = ENTITY_LABEL_KEYWORDS
        names = ENTITY_NAME_KEYWORDS
        label_hit = any(
            any(self._kw_match(kw, h) for h in headers) for kw in labels
        )
        name_hit = any(
            any(self._kw_match(kw, h) for h in headers) for kw in names
        )
        score = 0.0
        if label_hit:
            score += 0.5
        if name_hit:
            score += 0.5
        return score

    def _match_relationship_score(self, headers: list[str]) -> float:
        src_hit = any(
            any(self._kw_match(kw, h) for h in headers) for kw in REL_SOURCE_KEYWORDS
        )
        type_hit = any(
            any(self._kw_match(kw, h) for h in headers) for kw in REL_TYPE_KEYWORDS
        )
        tgt_hit = any(
            any(self._kw_match(kw, h) for h in headers) for kw in REL_TARGET_KEYWORDS
        )
        score = 0.0
        if src_hit:
            score += 0.34
        if type_hit:
            score += 0.34
        if tgt_hit:
            score += 0.32
        return score

    def _kw_match(self, keyword: str, header: str) -> bool:
        return keyword in header or header in keyword

    def _parse_sheet(self, ws, sheet_type: str) -> ParsedSheet:
        """解析指定 Sheet 为标准化数据。"""
        raw_headers = []
        row_iter = ws.iter_rows(min_row=1, max_row=1, values_only=True)
        first_row = next(row_iter, None)
        if first_row:
            raw_headers = [str(v).strip() if v is not None else "" for v in first_row]

        normalized = [self._strip_key(str(v)) if v else "" for v in raw_headers]

        # 构建列映射
        # 标准字段只映射第一次匹配，避免短关键词（如 "实体"）误匹配导致
        # 后续列（如 "实体描述"）覆盖前面的正确映射（如 "实体名称"）
        STANDARD_KEYS = {"label", "name", "source_name", "type", "target_name"}
        column_map = {}
        for idx, h in enumerate(normalized):
            mapped = self._map_column(h, sheet_type)
            if not mapped:
                continue
            if mapped in STANDARD_KEYS and mapped in column_map:
                # 标准字段已映射过，当前列降级为属性列（用原始列名作 key）
                column_map[h] = idx
            else:
                column_map[mapped] = idx

        # 读取数据行
        rows: list[dict] = []
        for row_idx, row_cells in enumerate(
            ws.iter_rows(min_row=2, values_only=True), start=2
        ):
            row_dict = {**{k: None for k in column_map}}
            for key, col_idx in column_map.items():
                val = row_cells[col_idx] if col_idx < len(row_cells) else None
                row_dict[key] = val
            # 只保留非空行
            if any(v is not None and str(v).strip() != "" for v in row_dict.values()):
                row_dict["_row"] = row_idx
                rows.append(row_dict)

        return ParsedSheet(
            sheet_name=ws.title,
            sheet_type=sheet_type,
            headers=raw_headers,
            column_map=column_map,
            rows=rows,
            row_count=len(rows),
        )

    def _parse_sheet_by_name(self, file_bytes: bytes, sheet_name: str, sheet_type: str) -> Optional[ParsedSheet]:
        try:
            wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        except Exception:
            return None

        if sheet_name not in wb.sheetnames:
            wb.close()
            return None

        result = self._parse_sheet(wb[sheet_name], sheet_type)
        wb.close()
        return result

    @staticmethod
    def _strip_key(raw: str) -> str:
        """归一化列名用于匹配。"""
        return raw.strip().lower().replace("_", "").replace(" ", "")

    def _map_column(self, normalized: str, sheet_type: str) -> Optional[str]:
        """将归一化后的列名映射到标准键。精确匹配优先于子串匹配。"""
        # 关键词也需归一化后比较
        def norm(kw: str) -> str:
            return self._strip_key(kw)

        if sheet_type == "entity":
            # 第一轮：精确匹配（避免短关键词子串误匹配）
            for kw in ENTITY_LABEL_KEYWORDS:
                if normalized == norm(kw):
                    return "label"
            for kw in ENTITY_NAME_KEYWORDS:
                if normalized == norm(kw):
                    return "name"
            # 第二轮：子串匹配（兜底）
            for kw in ENTITY_LABEL_KEYWORDS:
                if self._kw_match(kw, normalized):
                    return "label"
            for kw in ENTITY_NAME_KEYWORDS:
                if self._kw_match(kw, normalized):
                    return "name"
            # 其他列作为属性
            return normalized
        elif sheet_type == "relationship":
            # 第一轮：精确匹配
            for kw in REL_SOURCE_KEYWORDS:
                if normalized == norm(kw):
                    return "source_name"
            for kw in REL_TYPE_KEYWORDS:
                if normalized == norm(kw):
                    return "type"
            for kw in REL_TARGET_KEYWORDS:
                if normalized == norm(kw):
                    return "target_name"
            # 第二轮：子串匹配（兜底）
            for kw in REL_SOURCE_KEYWORDS:
                if self._kw_match(kw, normalized):
                    return "source_name"
            for kw in REL_TYPE_KEYWORDS:
                if self._kw_match(kw, normalized):
                    return "type"
            for kw in REL_TARGET_KEYWORDS:
                if self._kw_match(kw, normalized):
                    return "target_name"
            return normalized
        return None

"""数据导入引擎 —— 支持 Excel 和关系数据库导入（v3.0: SQLite + Neo4j prefix 模式）."""

import io

from backend.repositories.neo4j_repository import Neo4jRepository
from backend.services.cypher_generator import CypherBatch, CypherGenerator
from backend.services.excel_detector import ParsedSheet
from backend.services.neo4j_tx_manager import Neo4jTransactionManager, ExecuteResult
from backend.models.import_task import (
    DBConnection, TableInfo, ImportPreviewData, ImportResult,
    TableMapping, RelationshipMapping,
)
from backend.services.system_service import SystemService
from backend.models.system import UpsertRelationSemanticRequest


class ImportService:
    """数据导入引擎"""

    DB_DIALECTS = {
        "mysql":      "mysql+pymysql://{user}:{password}@{host}:{port}/{database}",
        "postgresql": "postgresql+psycopg2://{user}:{password}@{host}:{port}/{database}",
        "mssql":      "mssql+pyodbc://{user}:{password}@{host}:{port}/{database}?driver=ODBC+Driver+17+for+SQL+Server",
        "oracle":     "oracle+cx_oracle://{user}:{password}@{host}:{port}/{database}",
        "sqlite":     "sqlite:///{database}",
    }

    def __init__(self, repo: Neo4jRepository, system_service: SystemService):
        self.repo = repo
        self.system_service = system_service
        self.tx_manager = Neo4jTransactionManager(repo)

    # ==================== Excel 导入 ====================

    def preview_excel(self, file_bytes: bytes) -> ImportPreviewData:
        """解析 Excel 并返回预览数据（跳过"使用说明"Sheet）."""
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True)

        entities = []
        relationships = []

        # 跳过说明类 Sheet（以"使用"开头）
        data_sheets = [s for s in wb.sheetnames if not s.startswith("使用")]

        if "实体" in data_sheets:
            ws = wb["实体"]
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            headers = [cell.value for cell in ws[1]]
            for row in rows:
                if row[0] is None:
                    continue
                entity = {
                    "label": str(row[0]) if row[0] else "",
                    "name": str(row[1]) if len(row) > 1 and row[1] else "Unnamed",
                }
                for i, h in enumerate(headers):
                    if i >= 2 and h and row[i] is not None:
                        entity[h] = str(row[i])
                entities.append(entity)

        if "关系" in data_sheets:
            ws = wb["关系"]
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            for row in rows:
                if row[0] is None:
                    continue
                rel = {
                    "source_name": str(row[0]) if row[0] else "",
                    "type": str(row[1]) if len(row) > 1 and row[1] else "",
                    "target_name": str(row[2]) if len(row) > 2 and row[2] else "",
                }
                relationships.append(rel)

        wb.close()
        return ImportPreviewData(
            entities=entities,
            relationships=relationships,
            total_entities=len(entities),
            total_relationships=len(relationships),
        )

    def import_from_excel(
        self, file_bytes: bytes, system_name: str, description: str,
        prefix: str = "",
    ) -> ImportResult:
        """从 Excel 导入创建新系统（v3.5: 支持自定义 prefix）。"""
        # 1. 在 SQLite 中创建系统记录（prefix 为空则自动生成）
        system = self.system_service.create_system(
            name=system_name, description=description, import_source="excel",
            prefix=prefix,
        )
        prefix = system.prefix

        # 2. 解析 Excel
        preview = self.preview_excel(file_bytes)
        errors = []

        # 3. 批量写入 Neo4j（使用 prefix）
        entity_count = self.repo.batch_create_nodes(preview.entities, prefix)

        # 4. 批量写入关系（使用 prefix）
        rel_count = 0
        if preview.relationships:
            rel_count = self.repo.batch_create_relationships(preview.relationships, prefix)

        # 5. 更新 SQLite 统计
        node_total = self.repo.count_system_nodes(prefix)
        rel_total = self.repo.count_system_relationships(prefix)
        self.system_service.update_counts(system.system_id, node_total, rel_total)

        # 6. v3.6: 自动初始化关系语义
        self._auto_init_semantics(prefix)

        return ImportResult(
            success=len(errors) == 0,
            system_id=system.system_id,
            system_name=system.name,
            entities_created=entity_count,
            relationships_created=rel_count,
            message=f"成功导入 {entity_count} 个实体, {rel_count} 条关系",
            errors=errors,
        )

    def append_from_excel(
        self, file_bytes: bytes, target_system_id: str,
    ) -> ImportResult:
        """★ v3.5: 将 Excel 数据追加到已有系统（不创建新系统）。"""
        # 1. 查询目标系统
        system = self.system_service.get_system(target_system_id)
        if not system:
            return ImportResult(
                success=False,
                system_id=target_system_id,
                system_name="",
                message=f"目标系统 '{target_system_id}' 不存在",
            )
        prefix = system.prefix

        # 2. 解析 Excel
        preview = self.preview_excel(file_bytes)
        errors = []

        # 3. 批量写入 Neo4j（使用目标系统的 prefix）
        entity_count = self.repo.batch_create_nodes(preview.entities, prefix)

        # 4. 批量写入关系
        rel_count = 0
        if preview.relationships:
            rel_count = self.repo.batch_create_relationships(preview.relationships, prefix)

        # 5. 更新 SQLite 统计
        node_total = self.repo.count_system_nodes(prefix)
        rel_total = self.repo.count_system_relationships(prefix)
        self.system_service.update_counts(system.system_id, node_total, rel_total)

        # 6. v3.6: 自动初始化关系语义（追加模式也可能引入新关系类型）
        self._auto_init_semantics(prefix)

        return ImportResult(
            success=len(errors) == 0,
            system_id=system.system_id,
            system_name=system.name,
            entities_created=entity_count,
            relationships_created=rel_count,
            message=f"已向「{system.name}」追加 {entity_count} 个实体, {rel_count} 条关系",
            errors=errors,
        )

    # ==================== 关系数据库导入 ====================

    def _build_url(self, conn: DBConnection) -> str:
        """构建 SQLAlchemy 连接 URL，避免用户名或密码破坏 URL。"""
        from sqlalchemy.engine import URL
        if conn.db_type not in self.DB_DIALECTS:
            raise ValueError(f"不支持的数据库类型: {conn.db_type}")
        if conn.db_type == "sqlite":
            database = conn.database
            if database == ":memory:":
                return URL.create("sqlite", database=":memory:").render_as_string(hide_password=False)
            return URL.create("sqlite", database=database).render_as_string(hide_password=False)
        drivers = {
            "mysql": "mysql+pymysql", "postgresql": "postgresql+psycopg2",
            "mssql": "mssql+pyodbc", "oracle": "oracle+cx_oracle",
        }
        query = {"driver": "ODBC Driver 17 for SQL Server"} if conn.db_type == "mssql" else {}
        return URL.create(
            drivers[conn.db_type], username=conn.user, password=conn.password,
            host=conn.host, port=conn.port, database=conn.database, query=query,
        ).render_as_string(hide_password=False)

    def _create_db_engine(self, conn: DBConnection):
        """创建短生命周期只读查询引擎。"""
        from sqlalchemy import create_engine
        kwargs = {"pool_pre_ping": True}
        if conn.db_type in {"mysql", "postgresql"}:
            kwargs["connect_args"] = {"connect_timeout": 5}
        return create_engine(self._build_url(conn), **kwargs)

    def test_connection(self, conn: DBConnection) -> bool:
        """测试数据库连接。"""
        from sqlalchemy import text
        try:
            engine = self._create_db_engine(conn)
        except ModuleNotFoundError as exc:
            raise ValueError(self._driver_error_message(conn, exc)) from exc
        try:
            with engine.connect() as c:
                c.execute(text("SELECT 1"))
            return True
        except Exception:
            return False
        finally:
            engine.dispose()

    @staticmethod
    def _driver_error_message(conn: DBConnection, exc: ModuleNotFoundError) -> str:
        drivers = {
            "mysql": "pymysql",
            "postgresql": "psycopg2-binary",
            "mssql": "pyodbc",
            "oracle": "cx_Oracle",
        }
        package = drivers.get(conn.db_type, exc.name or "对应数据库驱动")
        return f"当前环境缺少 {conn.db_type} 数据库驱动，请先安装依赖：pip install {package}"

    def get_tables(self, conn: DBConnection) -> list[TableInfo]:
        """获取数据库所有表结构."""
        from sqlalchemy import inspect
        engine = self._create_db_engine(conn)
        try:
            inspector = inspect(engine)
            tables = []
            for name in inspector.get_table_names():
                cols = inspector.get_columns(name)
                tables.append(TableInfo(
                    name=name,
                    columns=[{"name": c["name"], "type": str(c["type"])} for c in cols],
                ))
            return tables
        finally:
            engine.dispose()

    def preview_db(
        self, conn: DBConnection,
        entity_mappings: list[TableMapping] | None = None,
        relationship_mappings: list[RelationshipMapping] | None = None,
        limit: int = 100,
    ) -> ImportPreviewData:
        """预览数据库中的 Ontology 和 Relationship 数据。"""
        engine = self._create_db_engine(conn)
        entities: list[dict] = []
        relationships: list[dict] = []
        try:
            tables = self.get_tables(conn)
            by_name = {table.name.casefold(): table for table in tables}
            entity_table_name = conn.entity_table_name.strip()
            relationship_table_name = conn.relationship_table_name.strip()
            if not entity_table_name or not relationship_table_name:
                raise ValueError("请指定实体表名称和关系表名称")
            entity_table = by_name.get(entity_table_name.casefold())
            relationship_table = by_name.get(relationship_table_name.casefold())
            missing = [
                name for name, table in (
                    (entity_table_name, entity_table),
                    (relationship_table_name, relationship_table),
                ) if table is None
            ]
            if missing:
                raise ValueError(f"未找到指定表：{', '.join(missing)}")

            ontology_map = self._infer_entity_mapping(entity_table, entity_mappings)
            relationship_map = self._infer_relationship_mapping(relationship_table, relationship_mappings)
            for row_number, row in enumerate(self._read_table(engine, entity_table.name, limit), start=1):
                name = row.get(ontology_map.source_column)
                if name is None or str(name).strip() == "":
                    continue
                raw_label = (
                    row.get(ontology_map.label_column)
                    if ontology_map.label_column
                    else ontology_map.label_value
                )
                if raw_label is None or str(raw_label).strip() == "":
                    raise ValueError(
                        f"实体表第 {row_number} 行缺少有效 Label。"
                        "请在映射中指定 label_column 或 target_label。"
                    )
                label = str(raw_label).strip()
                entity = {"label": label, "name": str(name)}
                excluded_columns = {ontology_map.source_column, ontology_map.label_column}
                entity.update({
                    k: str(v) for k, v in row.items()
                    if k not in excluded_columns and v is not None
                })
                entities.append(entity)
            for row in self._read_table(engine, relationship_table.name, limit):
                source = row.get(relationship_map.source_column)
                target = row.get(relationship_map.target_column)
                if source is None or target is None or str(source).strip() == "" or str(target).strip() == "":
                    continue
                rel_type = relationship_map.type_value or row.get(relationship_map.type_column, "RELATED_TO")
                relationships.append({"type": str(rel_type), "source_name": str(source), "target_name": str(target)})
        finally:
            engine.dispose()
        return ImportPreviewData(entities=entities, relationships=relationships, total_entities=len(entities), total_relationships=len(relationships))

    @staticmethod
    def _pick_column(columns: list[str], candidates: tuple[str, ...], role: str) -> str:
        normalized = {column.casefold().replace("_", "").replace("-", ""): column for column in columns}
        for candidate in candidates:
            value = normalized.get(candidate.casefold().replace("_", "").replace("-", ""))
            if value:
                return value
        raise ValueError(f"{role}缺少可识别字段，候选字段：{', '.join(candidates)}")

    @staticmethod
    def _resolve_requested_column(columns: list[str], requested: str | None) -> str | None:
        if not requested:
            return None
        normalized = requested.casefold().replace("_", "").replace("-", "")
        return next(
            (column for column in columns
             if column.casefold().replace("_", "").replace("-", "") == normalized),
            None,
        )

    def _infer_entity_mapping(self, table: TableInfo, mappings: list[TableMapping] | None):
        from types import SimpleNamespace
        columns = [column["name"] for column in table.columns]
        mapping = mappings[0] if mappings else None
        name_column = self._resolve_requested_column(columns, mapping.source_column if mapping else None) or self._pick_column(columns, ("name", "Name", "entity_name", "node_name", "名称", "实体名称"), "Ontology 实体名称")
        requested_label_column = self._resolve_requested_column(
            columns, mapping.label_column if mapping else None
        )
        if mapping and mapping.label_column and not requested_label_column:
            raise ValueError(f"实体 Label 字段不存在：{mapping.label_column}")
        label_column = requested_label_column or self._find_optional_column(
            columns,
            (
                "label", "Label", "class", "Class", "type", "entity_type",
                "node_type", "node_type_name", "标签", "实体标签", "实体类型",
                "类别", "分类", "人员类别", "组织类型", "岗位类别",
            ),
        )
        label_value = mapping.target_label.strip() if mapping and mapping.target_label else None
        if label_value and self._resolve_requested_column(columns, label_value):
            label_value = None
        return SimpleNamespace(source_column=name_column, label_column=label_column, label_value=label_value)

    @staticmethod
    def _find_optional_column(columns: list[str], candidates: tuple[str, ...]) -> str | None:
        normalized = {column.casefold().replace("_", "").replace("-", ""): column for column in columns}
        for candidate in candidates:
            if candidate.casefold().replace("_", "").replace("-", "") in normalized:
                return normalized[candidate.casefold().replace("_", "").replace("-", "")]
        return None

    def _infer_relationship_mapping(self, table: TableInfo, mappings: list[RelationshipMapping] | None):
        from types import SimpleNamespace
        columns = [column["name"] for column in table.columns]
        mapping = mappings[0] if mappings else None
        source_column = self._resolve_requested_column(columns, mapping.source_column if mapping else None) or self._pick_column(columns, ("source_name", "Source_Name", "source", "src", "from", "源实体"), "Relationship 源实体")
        target_column = self._resolve_requested_column(columns, mapping.target_column if mapping else None) or self._pick_column(columns, ("target_name", "Target_Name", "target", "dst", "to", "目标实体"), "Relationship 目标实体")
        type_column = self._find_optional_column(columns, ("type", "Type", "relation", "Relation", "relationship", "rel_type", "关系类型"))
        type_value = mapping.relationship_type if mapping and mapping.relationship_type and mapping.relationship_type not in columns else None
        if not type_column and not type_value:
            raise ValueError("Relationship 缺少关系类型字段，请在映射中指定 relationship_type")
        return SimpleNamespace(source_column=source_column, type_column=type_column, type_value=type_value, target_column=target_column)

    def import_from_db(
        self, conn: DBConnection,
        entity_mappings: list[TableMapping],
        system_name: str, description: str,
        prefix: str = "",
        relationship_mappings: list[RelationshipMapping] | None = None,
    ) -> ImportResult:
        """分阶段执行关系数据库导入，并在异常时清理新建的 SQLite 系统记录。"""
        before_ids = {item.system_id for item in self.system_service.get_all_systems()}
        stages = [
            {"key": "mapping", "label": "表与字段匹配", "status": "success"},
            {"key": "reading", "label": "读取关系数据库数据", "status": "processing"},
            {"key": "cypher", "label": "转换为 Cypher", "status": "pending"},
            {"key": "execution", "label": "执行 Cypher", "status": "pending"},
            {"key": "result", "label": "校验执行结果", "status": "pending"},
        ]
        try:
            result = self._import_from_db_unsafe(
                conn, entity_mappings, system_name, description, prefix, relationship_mappings
            )
            for stage in stages:
                stage["status"] = "success"
            result.stages = stages
            return result
        except Exception as exc:
            for item in self.system_service.get_all_systems():
                if item.system_id not in before_ids:
                    self.system_service.delete_system(item.system_id)
            stages[-1]["status"] = "error"
            stages[-1]["message"] = str(exc)
            return ImportResult(
                success=False,
                system_name=system_name,
                stages=stages,
                message=f"关系数据库导入失败，已清理 SQLite 临时记录：{exc}",
                errors=[str(exc)],
            )

    def _import_from_db_unsafe(
        self, conn: DBConnection,
        entity_mappings: list[TableMapping],
        system_name: str, description: str,
        prefix: str = "",
        relationship_mappings: list[RelationshipMapping] | None = None,
    ) -> ImportResult:
        """执行关系数据库导入核心逻辑；外层负责异常补偿。"""
        # 1. 在 SQLite 创建系统记录
        system = self.system_service.create_system(
            name=system_name, description=description,
            import_source=f"database ({conn.db_type})",
            prefix=prefix,
        )
        prefix = system.prefix

        # 2. 预览获取数据
        preview = self.preview_db(conn, entity_mappings, relationship_mappings, limit=500000)

        errors = []

        # 3. 统一转换为标准 CypherBatch，并在写入前强制创建快照
        generator = CypherGenerator(prefix)
        entity_sheet = ParsedSheet(
            sheet_name=conn.entity_table_name,
            sheet_type="entity",
            headers=list(preview.entities[0].keys()) if preview.entities else [],
            rows=preview.entities,
            row_count=len(preview.entities),
        ) if preview.entities else None
        relationship_sheet = ParsedSheet(
            sheet_name=conn.relationship_table_name,
            sheet_type="relationship",
            headers=list(preview.relationships[0].keys()) if preview.relationships else [],
            rows=preview.relationships,
            row_count=len(preview.relationships),
        ) if preview.relationships else None
        batch = generator.generate(entity_sheet, relationship_sheet, strategy="CREATE")
        execution = self.tx_manager.execute_with_backup(batch, prefix)
        if execution.success:
            node_total = self.repo.count_system_nodes(prefix)
            rel_total = self.repo.count_system_relationships(prefix)
            self.system_service.update_counts(system.system_id, node_total, rel_total)
        return ImportResult(
            success=execution.success,
            system_id=system.system_id,
            system_name=system.name,
            entities_created=execution.entities_created,
            relationships_created=execution.relationships_created,
            snapshot_id=execution.snapshot_id,
            backup_available=execution.backup_available,
            warnings=execution.warnings,
            message=execution.message or f"成功导入 {execution.entities_created} 个实体, {execution.relationships_created} 条关系",
            errors=execution.errors,
            stages=[
                {"key": "mapping", "label": "表与字段匹配", "status": "success"},
                {"key": "reading", "label": "读取关系数据库数据", "status": "success"},
                {"key": "cypher", "label": "转换为 Cypher", "status": "success"},
                {"key": "execution", "label": "执行 Cypher", "status": "success" if execution.success else "error"},
                {"key": "result", "label": "校验执行结果", "status": "success" if execution.success else "error"},
            ],
        )

    def append_from_db(
        self, conn: DBConnection,
        entity_mappings: list[TableMapping],
        target_system_id: str,
        relationship_mappings: list[RelationshipMapping] | None = None,
    ) -> ImportResult:
        """★ v3.5: 将数据库数据追加到已有系统（不创建新系统）。"""
        # 1. 查询目标系统
        system = self.system_service.get_system(target_system_id)
        if not system:
            return ImportResult(
                success=False,
                system_id=target_system_id,
                system_name="",
                message=f"目标系统 '{target_system_id}' 不存在",
            )
        prefix = system.prefix

        # 2. 预览获取数据
        preview = self.preview_db(conn, entity_mappings, relationship_mappings, limit=500000)

        errors = []

        # 3. 统一转换为标准 CypherBatch，并在写入前强制创建快照
        generator = CypherGenerator(prefix)
        entity_sheet = ParsedSheet(
            sheet_name=system.name,
            sheet_type="entity",
            headers=list(preview.entities[0].keys()) if preview.entities else [],
            rows=preview.entities,
            row_count=len(preview.entities),
        ) if preview.entities else None
        relationship_sheet = ParsedSheet(
            sheet_name=conn.relationship_table_name,
            sheet_type="relationship",
            headers=list(preview.relationships[0].keys()) if preview.relationships else [],
            rows=preview.relationships,
            row_count=len(preview.relationships),
        ) if preview.relationships else None
        batch = generator.generate(entity_sheet, relationship_sheet, strategy="MERGE")
        execution = self.tx_manager.execute_with_backup(batch, prefix)
        if execution.success:
            node_total = self.repo.count_system_nodes(prefix)
            rel_total = self.repo.count_system_relationships(prefix)
            self.system_service.update_counts(system.system_id, node_total, rel_total)
        return ImportResult(
            success=execution.success,
            system_id=system.system_id,
            system_name=system.name,
            entities_created=execution.entities_created,
            relationships_created=execution.relationships_created,
            snapshot_id=execution.snapshot_id,
            backup_available=execution.backup_available,
            warnings=execution.warnings,
            message=execution.message or f"已向「{system.name}」追加 {execution.entities_created} 个实体, {execution.relationships_created} 条关系",
            errors=execution.errors,
        )

    # ==================== 模板下载 ====================

    def generate_template(self) -> bytes:
        """生成 Excel 导入模板（含使用说明 + 样例数据）."""
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()

        # ═══════════════════════════════════════
        # 公共样式
        # ═══════════════════════════════════════
        header_font = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell_align = Alignment(vertical="top", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )
        title_font = Font(name="微软雅黑", bold=True, size=16, color="1F4E79")
        section_font = Font(name="微软雅黑", bold=True, size=12, color="2E75B6")
        body_font = Font(name="微软雅黑", size=10, color="333333")

        # ═══════════════════════════════════════
        # Sheet 1: 使用说明
        # ═══════════════════════════════════════
        ws_intro = wb.active
        ws_intro.title = "使用说明"
        ws_intro.sheet_properties.tabColor = "4472C4"
        ws_intro.column_dimensions["A"].width = 22
        ws_intro.column_dimensions["B"].width = 65

        intro_data = [
            ("📋 知识图谱导入模板说明", None, title_font),
            (None, None, None),
            ("一、模板概述", "本模板用于将结构化数据批量导入到知识图谱系统中。\n"
                           "请按照「实体」和「关系」两个 Sheet 的格式填写数据，\n"
                           "完成后通过系统右上角的「导入」功能上传即可。", section_font),
            (None, None, None),
            ("二、实体 Sheet 说明", "每行代表一个知识图谱节点（实体），包含以下列：", section_font),
            ("  ─ 标签(Label)", "必填。实体的类型，可参考下方「实体类型举例」，也可自定义。\n"
                              "如：Disease / Drug / Symptom / SideEffect / BodyPart", body_font),
            ("  ─ 名称(name)", "必填。实体的名称，同一类型下建议保持唯一。\n"
                              "如：糖尿病 / 二甲双胍 / 头晕", body_font),
            ("  ─ 描述(description)", "选填。对实体的详细描述说明。", body_font),
            ("  ─ 别名(alias)", "选填。实体的别名，多个别名用逗号分隔。", body_font),
            (None, None, None),
            ("三、关系 Sheet 说明", "每行代表两个实体之间的一条关系，包含以下列：", section_font),
            ("  ─ 源实体名称(source_name)", "必填。关系起点的实体名称，须与「实体」Sheet 中的名称一致。", body_font),
            ("  ─ 关系类型(type)", "必填。关系的类型，可参考下方「关系类型举例」，也可自定义。", body_font),
            ("  ─ 目标实体名称(target_name)", "必填。关系终点的实体名称，须与「实体」Sheet 中的名称一致。", body_font),
            (None, None, None),
            ("四、实体类型举例", None, section_font),
            (None, "Disease  → 疾病\nDrug  → 药物/药品\nSymptom  → 症状\nSideEffect  → 副作用\n"
                   "BodyPart  → 身体部位\n* 你可以自定义其他类型，系统会自动创建", body_font),
            (None, None, None),
            ("五、关系类型举例", None, section_font),
            (None, "TREATS  → 药物-治疗→疾病\nHAS_SYMPTOM  → 疾病-有症状→症状\n"
                   "HAS_SIDE_EFFECT  → 药物-有副作用→副作用\nAFFECTS  → 疾病/药物-影响→身体部位\n"
                   "BELONGS_TO  → A-属于→B\nSUBCLASS_OF  → A-是子类→B\n"
                   "INTERACTS_WITH  → A-与B相互作用\n* 你可以自定义其他关系类型，系统会自动创建",
             body_font),
            (None, None, None),
            ("六、使用步骤", None, section_font),
            (None, "1. 在「实体」Sheet 中填写所有要导入的节点数据\n"
                   "2. 在「关系」Sheet 中填写节点之间的关系\n"
                   "3. （可选）删除不需要的样例行\n"
                   "4. 点击系统右上角「导入」→「从Excel导入」上传文件\n"
                   "5. 预览确认后，设置系统名称并完成导入",
             body_font),
            (None, None, None),
            ("七、注意事项", None, section_font),
            (None, "• 请勿修改 Sheet 名称和列标题（首行）\n"
                   "• 实体名称和关系类型区分大小写\n"
                   "• 关系的源/目标实体名称必须在「实体」Sheet 中已定义\n"
                   "• 空白行会被自动忽略\n"
                   "• 本「使用说明」Sheet 在导入时会被自动跳过",
             body_font),
        ]

        for row_idx, (col_a, col_b, font) in enumerate(intro_data, start=1):
            ws_intro.row_dimensions[row_idx].height = 20
            if col_a is not None:
                cell_a = ws_intro.cell(row=row_idx, column=1, value=col_a)
                cell_a.font = font or body_font
                cell_a.alignment = cell_align
            if col_b is not None:
                cell_b = ws_intro.cell(row=row_idx, column=2, value=col_b)
                cell_b.font = font or body_font
                cell_b.alignment = cell_align
            # 标题行合并单元格
            if row_idx == 1:
                ws_intro.merge_cells("A1:B1")
                cell = ws_intro["A1"]
                cell.alignment = Alignment(horizontal="center", vertical="center")
                ws_intro.row_dimensions[1].height = 35

        # ═══════════════════════════════════════
        # Sheet 2: 实体
        # ═══════════════════════════════════════
        ws_entity = wb.create_sheet("实体")
        ws_entity.sheet_properties.tabColor = "70AD47"
        ws_entity.column_dimensions["A"].width = 16
        ws_entity.column_dimensions["B"].width = 20
        ws_entity.column_dimensions["C"].width = 40
        ws_entity.column_dimensions["D"].width = 25

        entity_headers = ["标签(Label)", "名称(name)", "描述(description)", "别名(alias)"]
        for col_idx, h in enumerate(entity_headers, start=1):
            cell = ws_entity.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        # 样例数据（疾病、药物、症状、副作用、身体部位）
        sample_entities = [
            ["Disease",    "糖尿病", "一种以高血糖为特征的代谢性疾病", "消渴症,DM"],
            ["Disease",    "高血压", "动脉血压持续升高的慢性疾病", "HPN"],
            ["Disease",    "冠心病", "冠状动脉粥样硬化导致的心脏病", "CHD,冠状动脉性心脏病"],
            ["Drug",       "二甲双胍", "口服降糖药，2型糖尿病一线用药", "Metformin"],
            ["Drug",       "硝苯地平", "钙通道阻滞剂，用于高血压和心绞痛", "Nifedipine"],
            ["Drug",       "阿司匹林", "抗血小板药物，预防心血管事件", "Aspirin"],
            ["Symptom",    "头晕", "感觉自身或周围环境旋转的不适感", "眩晕"],
            ["Symptom",    "乏力", "全身无力、容易疲劳的症状", "疲倦"],
            ["Symptom",    "多饮多尿", "饮水量和排尿量异常增多", "烦渴多尿"],
            ["SideEffect", "恶心", "上腹部不适、有呕吐欲望的感觉", ""],
            ["SideEffect", "胃肠道不适", "腹痛、腹泻、恶心等消化道反应", "肠胃不适"],
            ["BodyPart",   "心脏", "人体循环系统的核心器官", "心"],
            ["BodyPart",   "肝脏", "人体最大的消化腺和代谢器官", "肝"],
            ["BodyPart",   "肾脏", "人体的排泄和内分泌器官", "肾"],
        ]
        for row_idx, row_data in enumerate(sample_entities, start=2):
            ws_entity.row_dimensions[row_idx].height = 22
            for col_idx, val in enumerate(row_data, start=1):
                cell = ws_entity.cell(row=row_idx, column=col_idx, value=val)
                cell.font = body_font
                cell.alignment = cell_align
                cell.border = thin_border
            # 区分不同类型用不同背景色（仅提示作用）
            label_colors = {
                "Disease": "FCE4D6", "Drug": "D9E2F3", "Symptom": "E2EFDA",
                "SideEffect": "FFF2CC", "BodyPart": "F2DCDB",
            }
            fill_color = label_colors.get(row_data[0])
            if fill_color:
                for col_idx in range(1, 5):
                    ws_entity.cell(row=row_idx, column=col_idx).fill = PatternFill(
                        start_color=fill_color, end_color=fill_color, fill_type="solid"
                    )

        # ═══════════════════════════════════════
        # Sheet 3: 关系
        # ═══════════════════════════════════════
        ws_rel = wb.create_sheet("关系")
        ws_rel.sheet_properties.tabColor = "ED7D31"
        ws_rel.column_dimensions["A"].width = 22
        ws_rel.column_dimensions["B"].width = 24
        ws_rel.column_dimensions["C"].width = 22

        rel_headers = ["源实体名称(source_name)", "关系类型(type)", "目标实体名称(target_name)"]
        for col_idx, h in enumerate(rel_headers, start=1):
            cell = ws_rel.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        # 样例关系数据
        sample_relationships = [
            ["二甲双胍",  "TREATS",           "糖尿病"],
            ["硝苯地平",  "TREATS",           "高血压"],
            ["阿司匹林",  "TREATS",           "冠心病"],
            ["糖尿病",    "HAS_SYMPTOM",      "头晕"],
            ["糖尿病",    "HAS_SYMPTOM",      "乏力"],
            ["糖尿病",    "HAS_SYMPTOM",      "多饮多尿"],
            ["高血压",    "HAS_SYMPTOM",      "头晕"],
            ["二甲双胍",  "HAS_SIDE_EFFECT",  "恶心"],
            ["二甲双胍",  "HAS_SIDE_EFFECT",  "胃肠道不适"],
            ["阿司匹林",  "HAS_SIDE_EFFECT",  "胃肠道不适"],
            ["硝苯地平",  "HAS_SIDE_EFFECT",  "头晕"],
            ["糖尿病",    "AFFECTS",          "心脏"],
            ["糖尿病",    "AFFECTS",          "肾脏"],
            ["冠心病",    "AFFECTS",          "心脏"],
            ["高血压",    "AFFECTS",          "心脏"],
        ]
        for row_idx, row_data in enumerate(sample_relationships, start=2):
            ws_rel.row_dimensions[row_idx].height = 22
            for col_idx, val in enumerate(row_data, start=1):
                cell = ws_rel.cell(row=row_idx, column=col_idx, value=val)
                cell.font = body_font
                cell.alignment = cell_align
                cell.border = thin_border
            # 按关系类型着色
            rel_colors = {
                "TREATS": "E2EFDA", "HAS_SYMPTOM": "D9E2F3",
                "HAS_SIDE_EFFECT": "FFF2CC", "AFFECTS": "F2DCDB",
            }
            fill_color = rel_colors.get(row_data[1])
            if fill_color:
                for col_idx in range(1, 4):
                    ws_rel.cell(row=row_idx, column=col_idx).fill = PatternFill(
                        start_color=fill_color, end_color=fill_color, fill_type="solid"
                    )

        # ── 冻结首行（表头始终可见） ──
        ws_entity.freeze_panes = "A2"
        ws_rel.freeze_panes = "A2"

        # ── 输出 ──
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    def _read_table(self, engine, table_name: str, limit: int) -> list[dict]:
        """通过 SQLAlchemy 元数据读取表数据，兼容不同方言并避免拼接表名。"""
        from sqlalchemy import MetaData, Table, select
        if not table_name or limit <= 0:
            return []
        table = Table(table_name, MetaData(), autoload_with=engine)
        with engine.connect() as connection:
            result = connection.execute(select(table).limit(min(limit, 500000)))
            return [dict(row._mapping) for row in result]

    # ═══════════════════════════════════════════
    # v3.6: 语义自动初始化
    # ═══════════════════════════════════════════

    def _auto_init_semantics(self, prefix: str) -> None:
        """导入完成后自动从 Neo4j 扫描关系类型并初始化语义（不覆盖已有配置）。"""
        try:
            rel_types = self.repo.get_relation_types_by_prefix(prefix)
            if not rel_types:
                return
            from backend.services.query_service import get_preset_semantics
            # 先用预置语义填充已知类型
            for rt in rel_types:
                preset = get_preset_semantics(rt)
                if preset:
                    self.system_service.upsert_relation_semantic(
                        prefix,
                        UpsertRelationSemanticRequest(**preset, rel_type=rt),
                    )
            # 初始化其余未配置的类型
            count = self.system_service.init_semantics_from_neo4j(prefix, rel_types)
            if count > 0:
                import logging
                logger = logging.getLogger(__name__)
                logger.info(f"v3.6: 自动初始化 {count} 条关系语义 (prefix={prefix})")
        except Exception:
            import logging
            logger = logging.getLogger(__name__)
            logger.warning(f"v3.6: 自动初始化语义失败，可忽略 (prefix={prefix})")
